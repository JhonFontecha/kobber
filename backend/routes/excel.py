from datetime import datetime
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel

from database import get_client

router = APIRouter()


class MLVariantPct(BaseModel):
    variant_id: str
    porcentaje: Optional[float] = None


class MLExportItem(BaseModel):
    product_id: str
    variantes: list[MLVariantPct] = []


class MLExportRequest(BaseModel):
    items: list[MLExportItem]


@router.post("/generate-ml")
def generate_ml_excel(body: MLExportRequest):
    if not body.items:
        raise HTTPException(400, "No hay productos para exportar")

    db = get_client()
    ids = [i.product_id for i in body.items]
    # Mapa variant_id -> porcentaje
    variant_pct_map: dict = {}
    for item in body.items:
        for v in item.variantes:
            variant_pct_map[v.variant_id] = v.porcentaje

    products = db.table("products").select(
        "id, nombre, descripcion, marca, categoria, subcategoria, caracteristicas, "
        "product_variants(id, clave, codigo, descripcion, precio_distribuidor, nc, unidades_caja, stock), "
        "product_images(url, orden)"
    ).in_("id", ids).execute().data

    if not products:
        raise HTTPException(404, "No se encontraron productos")

    wb = Workbook()
    ws = wb.active
    ws.title = "Publicar ML"

    # ── Estilos ────────────────────────────────────────────────────────────────
    COPPER   = "C8762C"
    WHITE    = "FFFFFF"
    ALT_ROW  = "FDF1E4"
    BORDER_C = "D4CEC5"

    thin   = Side(style="thin",   color=BORDER_C)
    medium = Side(style="medium", color=BORDER_C)
    border_data   = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_header = Border(left=medium, right=medium, top=medium, bottom=medium)

    h_font   = Font(bold=True, color=WHITE,  size=10, name="Calibri")
    d_font   = Font(color="1A1510",          size=10, name="Calibri")
    price_font = Font(color="1A1510",        size=10, name="Calibri", bold=True)
    green_font = Font(color="2E7D52",        size=10, name="Calibri", bold=True)
    h_fill   = PatternFill("solid", fgColor=COPPER)
    alt_fill = PatternFill("solid", fgColor=ALT_ROW)
    center   = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left     = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── Columnas ───────────────────────────────────────────────────────────────
    headers = [
        ("Título (max 60 c.)",    45),
        ("SKU / Clave",           16),
        ("Código Trupper",        15),
        ("EAN / NC",              14),
        ("Condición",             10),
        ("Marca",                 12),
        ("Modelo",                16),
        ("Precio [$]",            17),
        ("Stock",                  7),
        ("Uds/Caja",               9),
        ("Descripción",           55),
        ("Categoría",             30),
        ("Forma de envío",        16),
        ("Cuotas",                10),
        ("Retiro en persona",     16),
        ("Garantía",              12),
        ("Fotos (URLs)",          60),
    ]

    # Fila de encabezado
    ws.row_dimensions[1].height = 24
    for col, (title, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font      = h_font
        cell.fill      = h_fill
        cell.border    = border_header
        cell.alignment = center
        ws.column_dimensions[cell.column_letter].width = width

    # ── Filas de datos ─────────────────────────────────────────────────────────
    row_idx = 2
    id_order = {pid: i for i, pid in enumerate(ids)}
    products.sort(key=lambda p: id_order.get(p["id"], 999))

    for p in products:
        fotos      = sorted(p.get("product_images") or [], key=lambda x: x.get("orden", 0))
        fotos_str  = ",".join(f["url"] for f in fotos)   # ML requiere coma como separador
        variantes  = p.get("product_variants") or []
        is_alt     = (row_idx % 2 == 0)
        fill       = alt_fill if is_alt else PatternFill()

        if not variantes:
            variantes = [{}]

        for v in variantes:
            precio_base  = v.get("precio_distribuidor")
            pct          = variant_pct_map.get(v.get("id"))
            precio_venta = round(precio_base * (1 + (pct or 0) / 100)) if precio_base else None

            # Título: nombre + clave, truncado a 60 caracteres
            clave       = v.get("clave") or ""
            titulo_raw  = f"{p.get('nombre', '')} {p.get('marca', '')} {clave}".strip()
            titulo      = titulo_raw[:60]

            row_data = [
                titulo,
                clave,
                str(v.get("codigo") or ""),
                str(v.get("nc") or ""),
                "Nuevo",
                p.get("marca") or "",
                clave,
                precio_venta,
                v.get("stock") or 0,
                v.get("unidades_caja") or "",
                p.get("descripcion") or "",
                p.get("categoria") or "",
                "Mercado Envíos",
                "Cuotas",
                "Seleccionar",
                "Seleccionar",
                fotos_str,
            ]

            ws.row_dimensions[row_idx].height = 18
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border    = border_data
                cell.fill      = fill
                if col == 8:        # precio final
                    cell.font      = green_font
                    cell.alignment = center
                elif col in (9, 10):  # stock, uds/caja
                    cell.font      = d_font
                    cell.alignment = center
                elif col == 11:     # descripción larga
                    cell.font      = d_font
                    cell.alignment = left
                else:
                    cell.font      = d_font
                    cell.alignment = left

            row_idx += 1

    # Freeze header
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"kobber_ML_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


class ExportRequest(BaseModel):
    product_ids:    Optional[list[str]] = None   # None = todos
    margen_global:  Optional[float]     = None   # % de margen sobre precio distribuidor
    categoria_ml:   Optional[str]       = None   # sobreescribe categoría si se especifica


def _precio_venta(precio_base: float, margen: float) -> float:
    return round(precio_base * (1 + margen / 100), 2)


@router.post("/export")
def export_excel(body: ExportRequest):
    db = get_client()

    # Obtener productos con variantes y atributos
    query = db.table("products").select(
        "id, nombre, descripcion, marca, categoria, subcategoria, seccion, caracteristicas, estado, pagina_catalogo, "
        "product_attributes(nombre, valor, unidad, variant_id), "
        "product_variants(id, codigo, clave, descripcion, precio_distribuidor, nc, "
        "unidades_caja, unidades_master, stock, estado, "
        "product_attributes(nombre, valor, unidad)), "
        "product_images(url, orden)"
    )

    if body.product_ids:
        query = query.in_("id", body.product_ids)
    else:
        query = query.neq("estado", "descartado")

    products = query.order("categoria").order("nombre").execute().data

    if not products:
        raise HTTPException(status_code=404, detail="No hay productos para exportar")

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    headers = [
        "Producto ID",
        "Nombre producto",
        "Descripción producto",
        "Marca",
        "Categoría",
        "Subcategoría",
        "Sección",
        "Características",
        "Atributos familia",
        "Variante ID",
        "Código pedido",
        "Clave / SKU",
        "Descripción variante",
        "Precio distribuidor",
        "Precio venta (con margen)",
        "Margen %",
        "NC",
        "Uds/caja",
        "Uds/máster",
        "Stock",
        "Estado variante",
        "Atributos variante",
        "Fotos (URLs)",
        "Estado producto",
        "Página catálogo",
    ]

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    def _attrs_str(attrs):
        return " | ".join(
            f"{a['nombre']}: {a['valor']}{' ' + a['unidad'] if a.get('unidad') else ''}"
            for a in (attrs or [])
        )

    # Una fila por variante
    row_idx = 2
    for p in products:
        fotos = sorted(p.get("product_images") or [], key=lambda x: x.get("orden", 0))
        fotos_str = ",".join(f["url"] for f in fotos)   # ML requiere coma como separador

        caracteristicas_str = " | ".join(p.get("caracteristicas") or [])

        # Atributos de familia (variant_id is None)
        familia_attrs = [a for a in (p.get("product_attributes") or []) if not a.get("variant_id")]
        familia_attrs_str = _attrs_str(familia_attrs)

        variantes = p.get("product_variants") or []
        if not variantes:
            ws.append([
                p["id"], p.get("nombre") or "", p.get("descripcion") or "",
                p.get("marca") or "", p.get("categoria") or "",
                p.get("subcategoria") or "", p.get("seccion") or "",
                caracteristicas_str, familia_attrs_str,
                "", "", "", "", None, None, None, None, None, None, 0,
                "", "", fotos_str, p.get("estado") or "", p.get("pagina_catalogo"),
            ])
            row_idx += 1
            continue

        for v in variantes:
            margen = body.margen_global if body.margen_global is not None else 0
            precio_base = v.get("precio_distribuidor")
            precio_venta = _precio_venta(precio_base, margen) if precio_base else None

            ws.append([
                p["id"],
                p.get("nombre") or "",
                p.get("descripcion") or "",
                p.get("marca") or "",
                p.get("categoria") or "",
                p.get("subcategoria") or "",
                p.get("seccion") or "",
                caracteristicas_str,
                familia_attrs_str,
                v["id"],
                v.get("codigo") or "",
                v.get("clave") or "",
                v.get("descripcion") or "",
                precio_base,
                precio_venta,
                margen,
                v.get("nc"),
                v.get("unidades_caja"),
                v.get("unidades_master"),
                v.get("stock") or 0,
                v.get("estado") or "",
                _attrs_str(v.get("product_attributes")),
                fotos_str,
                p.get("estado") or "",
                p.get("pagina_catalogo"),
            ])
            row_idx += 1

    col_widths = [38, 45, 55, 12, 28, 22, 8, 50, 50, 38, 14, 16, 35, 18, 20, 10, 6, 8, 10, 8, 14, 50, 60, 14, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"kobber_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )

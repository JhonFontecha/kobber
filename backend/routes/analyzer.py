import glob
import io
import json
import os
import re
import subprocess
from collections import defaultdict
from typing import Optional

import anthropic
import openpyxl
from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from config import ANTHROPIC_API_KEY
from database import get_client

router = APIRouter()

# Palabras que ML pone en "Modelo" pero que NO son claves de producto
_INVALID_MODELO = {
    "seleccionar", "escribe o elige un valor", "",
    "truper", "pretul", "lusqtoff", "lüsqtoff", "urrea", "nuevo", "nuevo ",
}


def _extract_codes_from_title(title: str) -> list[str]:
    """
    Extrae posibles códigos del título cuando el campo Modelo no sirve.
    Busca números de 4-6 dígitos (códigos Trupper como 19861)
    y claves alfanuméricas tipo PR-108, MAN-15X.
    """
    nums   = re.findall(r'\b\d{4,6}\b', title)
    claves = re.findall(r'\b[A-Z]{2,6}-[\w/\'"]+\b', title, re.IGNORECASE)
    return [c.lower() for c in nums + claves]


def _find_match(
    row: dict,
    kobber_index: dict,
    bd_variants: dict,
) -> tuple[str | None, dict | None]:
    """
    Intenta encontrar la variante de BD que corresponde a una fila del archivo ML.
    Estrategias en orden:
      1. Por modelo (si no es marca/placeholder)
      2. Por SKU
      3. Partes del modelo (cuando ML combina claves: "A y B")
      4. Códigos extraídos del título (número Trupper embebido)
    Devuelve (clave_matched, bd_info) o (None, None).
    """
    modelo = row.get("modelo", "").strip()
    sku    = row.get("sku",    "").strip()
    titulo = row.get("titulo", "")

    candidatos: list[str] = []

    # 1 y 2: modelo y sku directos (filtrando marcas y placeholders)
    for val in [modelo, sku]:
        val_l = val.lower()
        if val_l and val_l not in _INVALID_MODELO:
            candidatos.append(val_l)

    # 3: modelo compuesto "CLAVE1 y CLAVE2"
    for sep in [" y ", " / ", ",", ";"]:
        if sep in modelo:
            for part in modelo.split(sep):
                part_l = part.strip().lower()
                if part_l and part_l not in _INVALID_MODELO:
                    candidatos.append(part_l)

    # 4: extraer códigos del título como último recurso
    candidatos += _extract_codes_from_title(titulo)

    for c in candidatos:
        if c in kobber_index:
            return c, bd_variants.get(c)
        if c in bd_variants:
            return c, bd_variants[c]

    return None, None

ISSUE_LABELS = {
    "NO_ENCONTRADO":    "No encontrado en BD",
    "SIN_PRECIO":       "Sin precio en ML",
    "PRECIO_DIFERENTE": "Precio muy diferente al de BD",
    "FALTA_EAN":        "BD tiene NC pero ML no tiene EAN",
    "SIN_EAN":          "Sin EAN ni NC en BD",
    "TITULO_LARGO":     "Título supera 60 caracteres",
    "SIN_CODIGO_ML":    "Sin código de catálogo ML (publicación libre)",
    "SIN_STOCK":        "Sin stock",
    "SIN_DESC":         "Sin descripción (BD tiene una)",
    "NO_EN_OUTPUT":     "Variante enviada a ML pero no aparece en el output",
}


def _detect_col(headers: list, keywords: list[str]) -> int | None:
    """Devuelve el índice (0-based) de la primera columna cuyo header contenga alguna keyword."""
    for i, h in enumerate(headers):
        h_low = str(h or "").lower()
        if any(k in h_low for k in keywords):
            return i
    return None


def _parse_kobber_file(content: bytes) -> dict:
    """
    Parsea el Excel generado por Kobber.
    Devuelve {clave_lower: {clave, variant_id, nombre, precio_dist, precio_venta}}.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "El archivo de Kobber está vacío")

    headers = [str(h or "").lower() for h in rows[0]]
    col_clave   = _detect_col(headers, ["clave", "sku", "modelo"])
    col_vid     = _detect_col(headers, ["variante id", "variant"])
    col_nombre  = _detect_col(headers, ["nombre producto", "nombre"])
    col_pdist   = _detect_col(headers, ["precio distribuidor", "p. dist"])
    col_pventa  = _detect_col(headers, ["precio venta", "p. venta"])

    if col_clave is None:
        raise HTTPException(400, "No se encontró columna de Clave/SKU en el archivo de Kobber")

    index: dict = {}
    for row in rows[1:]:
        clave = str(row[col_clave] or "").strip()
        if not clave:
            continue
        index[clave.lower()] = {
            "clave":       clave,
            "variant_id":  str(row[col_vid] or "")   if col_vid   is not None else "",
            "nombre":      str(row[col_nombre] or "") if col_nombre is not None else "",
            "precio_dist": row[col_pdist]             if col_pdist  is not None else None,
            "precio_venta":row[col_pventa]            if col_pventa is not None else None,
        }
    return index


_ML_FIELD_KEYWORDS: dict[str, list[str]] = {
    "codigo_ml":   ["catálogo ml", "catalogo ml"],
    "titulo":      ["título"],
    "ean":         ["código universal", "codigo universal"],
    "fotos":       ["fotos"],
    "sku":         ["sku"],
    "stock":       ["stock"],
    "precio":      ["precio [$]", "precio"],
    "descripcion": ["descripción", "descripcion"],
    "condicion":   ["condición", "condicion"],
    "marca":       ["marca"],
    "modelo":      ["modelo"],
}


def _ml_col_map(ws) -> dict[str, int]:
    """
    Detecta la fila de encabezados escaneando filas 2-5 y eligiendo
    la que más coincidencias tenga. Soporta templates antiguos (headers
    en fila 4) y nuevos (fila 3).
    """
    best_map:  dict[str, int] = {}
    best_hits: int = 0

    for row in range(2, 6):
        col_map: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            header = str(ws.cell(row, col).value or "").lower().strip()
            if not header:
                continue
            for field, keywords in _ML_FIELD_KEYWORDS.items():
                if field not in col_map and any(kw in header for kw in keywords):
                    col_map[field] = col
                    break
        if len(col_map) > best_hits:
            best_hits = len(col_map)
            best_map  = col_map

    return best_map


def _parse_ml_file(content: bytes) -> list[dict]:
    """
    Parsea el Excel generado por ML.
    Detecta columnas dinámicamente desde la fila 4 porque cada categoría
    puede tener un layout diferente. Devuelve filas con datos reales (fila >= 9).
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    hojas = [s for s in wb.sheetnames if s.lower() not in ["ayuda", "extra info"]]

    rows_out = []
    for hoja in hojas:
        ws   = wb[hoja]
        cols = _ml_col_map(ws)

        titulo_col = cols.get("titulo")
        if titulo_col is None:
            continue  # hoja sin columna de título reconocible

        def _cell(row_num: int, field: str, default=None):
            c = cols.get(field)
            if c is None:
                return default
            return ws.cell(row_num, c).value

        for row_num in range(9, ws.max_row + 1):
            titulo = str(ws.cell(row_num, titulo_col).value or "").strip()
            if not titulo:
                continue
            rows_out.append({
                "hoja":        hoja,
                "titulo":      titulo,
                "codigo_ml":   str(_cell(row_num, "codigo_ml") or "").strip(),
                "ean":         str(_cell(row_num, "ean")        or "").strip(),
                "sku":         str(_cell(row_num, "sku")        or "").strip(),
                "stock":       _cell(row_num, "stock"),
                "precio":      _cell(row_num, "precio"),
                "descripcion": str(_cell(row_num, "descripcion") or "").strip(),
                "condicion":   str(_cell(row_num, "condicion")   or "").strip(),
                "marca":       str(_cell(row_num, "marca")       or "").strip(),
                "modelo":      str(_cell(row_num, "modelo")      or "").strip(),
            })
    return rows_out


@router.post("/compare")
async def compare(
    kobber_file: UploadFile = File(...),
    ml_file:     UploadFile = File(...),
):
    kobber_bytes = await kobber_file.read()
    ml_bytes     = await ml_file.read()

    try:
        kobber_index = _parse_kobber_file(kobber_bytes)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Error leyendo archivo Kobber: {e}")

    try:
        ml_rows = _parse_ml_file(ml_bytes)
    except Exception as e:
        raise HTTPException(400, f"Error leyendo archivo ML: {e}")

    # Cargar variantes de BD para complementar info
    db = get_client()
    bd_variants: dict = {}   # clave_lower o codigo_lower → info
    for p in db.table("products").select(
        "nombre, descripcion, product_variants(id, clave, codigo, nc, precio_distribuidor, stock)"
    ).execute().data:
        for v in (p.get("product_variants") or []):
            info = {
                "nombre":      p.get("nombre", ""),
                "descripcion": p.get("descripcion", ""),
                "nc":          str(v.get("nc") or "").strip(),
                "precio_dist": v.get("precio_distribuidor"),
                "stock":       v.get("stock"),
            }
            clave  = str(v.get("clave")  or "").strip().lower()
            codigo = str(v.get("codigo") or "").strip().lower()
            if clave:  bd_variants[clave]  = info
            if codigo: bd_variants[codigo] = info

    # Rastrear qué claves del kobber aparecen en el ML output
    claves_en_ml: set = set()
    resultados = []

    for row in ml_rows:
        matched_clave, _ = _find_match(row, kobber_index, bd_variants)

        if matched_clave:
            claves_en_ml.add(matched_clave)

        bd_info = bd_variants.get(matched_clave) if matched_clave else None
        issues  = []

        if not matched_clave:
            issues.append(("NO_ENCONTRADO",
                f'Modelo "{row["modelo"]}" no coincide con ninguna clave de BD'))
        else:
            # Precio — el archivo ML ya lleva precio de venta (con margen aplicado),
            # solo avisamos si está vacío
            precio_ml = row["precio"]
            if precio_ml is None:
                issues.append(("SIN_PRECIO", "Precio no llenado en ML"))

            # EAN / NC
            ean = row["ean"]
            nc_bd = bd_info.get("nc", "") if bd_info else ""
            ean_invalido = not ean or ean.lower() in ("escribe o elige un valor", "")
            if ean_invalido:
                if nc_bd and nc_bd not in ("", "2"):
                    issues.append(("FALTA_EAN", f"BD tiene NC={nc_bd} — agrégalo como EAN"))
                else:
                    issues.append(("SIN_EAN", "Sin EAN en ML ni NC útil en BD"))

            # Título largo
            if len(row["titulo"]) > 60:
                issues.append(("TITULO_LARGO", f"{len(row['titulo'])} chars (máx 60)"))

            # Sin código catálogo ML
            if not row["codigo_ml"]:
                issues.append(("SIN_CODIGO_ML", "Publicación libre"))

            # Sin stock
            if row["stock"] is None:
                issues.append(("SIN_STOCK", "Stock no llenado en ML"))

            # Sin descripción
            if not row["descripcion"]:
                desc_bd = bd_info.get("descripcion", "") if bd_info else ""
                if desc_bd:
                    issues.append(("SIN_DESC", "ML sin descripción — BD tiene una"))

        resultados.append({
            "hoja":        row["hoja"],
            "titulo":      row["titulo"],
            "modelo":      row["modelo"],
            "bd_nombre":   bd_info.get("nombre", "") if bd_info else "",
            "matched":     matched_clave is not None,
            "codigo_ml":   row["codigo_ml"],
            "precio_ml":   row["precio"],
            "precio_bd":   bd_info.get("precio_dist") if bd_info else None,
            "issues":      [{"tipo": t, "mensaje": m} for t, m in issues],
        })

    # Claves enviadas a ML pero que no aparecen en el output
    no_en_output = []
    for clave_l, entry in kobber_index.items():
        if clave_l not in claves_en_ml:
            no_en_output.append({
                "clave":  entry["clave"],
                "nombre": entry["nombre"],
                "issues": [{"tipo": "NO_EN_OUTPUT",
                            "mensaje": "Esta variante fue enviada a ML pero no aparece en el output"}],
            })

    # Resumen
    tipo_cnt: dict = defaultdict(int)
    for r in resultados:
        for iss in r["issues"]:
            tipo_cnt[iss["tipo"]] += 1
    for r in no_en_output:
        tipo_cnt["NO_EN_OUTPUT"] += 1

    total_issues = sum(1 for r in resultados if r["issues"])

    return {
        "resumen": {
            "total_ml":      len(ml_rows),
            "total_kobber":  len(kobber_index),
            "con_problemas": total_issues,
            "sin_problemas": len(resultados) - total_issues,
            "no_en_output":  len(no_en_output),
            "por_tipo":      dict(tipo_cnt),
        },
        "resultados":   resultados,
        "no_en_output": no_en_output,
        "issue_labels": ISSUE_LABELS,
    }


# ── Generador de variaciones de título ────────────────────────────────────────

VARIATION_PROMPT = """\
Eres experto en nomenclatura de productos de ferretería y herramientas para e-commerce \
en Colombia y México (MercadoLibre).

Para cada producto de la lista genera EXACTAMENTE 3 títulos alternativos que:
1. Usen sinónimos o formas distintas de llamar el producto en Colombia/México \
   (ej: "pinzas" → "alicates", "juego" → "kit" o "set", "pistola" → "rociador", \
   "llave" → "llave inglesa" o "ajustable", "cautín" → "soldador de estaño", etc.)
2. Varíen la estructura del título (orden, adjetivos, contexto de uso)
3. Sean términos que los compradores realmente buscarían
4. Terminen SIEMPRE con el código del producto al final (ya incluido en el campo "codigo")
5. Tengan MÁXIMO 60 caracteres cada uno

Productos (JSON):
{productos_json}

Responde ÚNICAMENTE con un JSON array. Cada elemento es un array de 3 strings (los títulos).
Sin explicaciones, sin markdown. Solo el JSON.
Ejemplo de formato: [["título1a COD","título1b COD","título1c COD"],["título2a COD","título2b COD","título2c COD"]]"""


def _generate_variations_batch(client: anthropic.Anthropic, batch: list[dict]) -> list[list[str]]:
    """Llama a Claude para generar 3 variaciones de título por producto en el batch."""
    productos_input = [
        {"titulo": r["titulo"], "marca": r["marca"], "modelo": r["modelo"],
         "hoja": r["hoja"], "codigo": r["modelo"] or r["sku"]}
        for r in batch
    ]
    prompt = VARIATION_PROMPT.format(productos_json=json.dumps(productos_input, ensure_ascii=False))

    msg = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=2000,
        messages=[{"role": "user", "content": prompt}],
    )
    raw = msg.content[0].text.strip()

    # Limpiar posible markdown
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]

    parsed = json.loads(raw)
    # Garantizar que tenemos una lista del mismo largo que el batch
    result = []
    for i, item in enumerate(parsed[:len(batch)]):
        if isinstance(item, list):
            vars3 = [str(v)[:60] for v in item[:3]]
            while len(vars3) < 3:
                vars3.append(batch[i]["titulo"][:60])
            result.append(vars3)
        else:
            result.append([batch[i]["titulo"][:60]] * 3)
    return result


@router.post("/generate-variations")
async def generate_variations(ml_file: UploadFile = File(...)):
    """
    Recibe el archivo ML, genera 3 variaciones de título por fila usando Claude
    y devuelve un Excel con 3× las filas listas para subir a ML.
    """
    ml_bytes = await ml_file.read()
    try:
        ml_rows = _parse_ml_file(ml_bytes)
    except Exception as e:
        raise HTTPException(400, f"Error leyendo archivo ML: {e}")

    if not ml_rows:
        raise HTTPException(400, "No se encontraron productos en el archivo ML")

    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    # Procesar en batches de 10
    BATCH = 10
    all_variations: list[list[str]] = []
    for i in range(0, len(ml_rows), BATCH):
        batch = ml_rows[i:i + BATCH]
        try:
            vars_batch = _generate_variations_batch(client, batch)
        except Exception:
            # Si Claude falla en un batch, repetir el título original
            vars_batch = [[r["titulo"][:60]] * 3 for r in batch]
        all_variations.extend(vars_batch)

    # ── Construir Excel de salida ──────────────────────────────────────────────
    from datetime import datetime
    from io import BytesIO

    COPPER  = "C8762C"
    WHITE   = "FFFFFF"
    thin    = Side(style="thin", color="D4CEC5")
    border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_font  = Font(bold=True, color=WHITE, size=10, name="Calibri")
    d_font  = Font(color="1A1510",         size=10, name="Calibri")
    h_fill  = PatternFill("solid", fgColor=COPPER)
    fills   = [PatternFill(), PatternFill("solid", fgColor="FDF1E4")]
    center  = Alignment(horizontal="center", vertical="center")
    left    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Variaciones ML"

    headers = [
        ("Título",          50), ("Hoja ML",       22), ("Modelo / Clave", 16),
        ("Marca",           14), ("Precio [$]",     14), ("EAN",            14),
        ("Stock",            8), ("Condición",      10), ("Variación #",     5),
    ]
    ws.row_dimensions[1].height = 22
    for col, (h, w) in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = h_font; c.fill = h_fill; c.border = border; c.alignment = center
        ws.column_dimensions[c.column_letter].width = w

    row_idx = 2
    for row, variations in zip(ml_rows, all_variations):
        for var_num, titulo in enumerate(variations, 1):
            fill = fills[row_idx % 2]
            data = [
                titulo,
                row["hoja"],
                row["modelo"] or row["sku"],
                row["marca"],
                row["precio"],
                row["ean"] if row["ean"] not in ("", "Escribe o elige un valor") else "",
                row["stock"],
                row.get("condicion") or "Nuevo",
                var_num,
            ]
            ws.row_dimensions[row_idx].height = 16
            for col, val in enumerate(data, 1):
                c = ws.cell(row_idx, col, val)
                c.border = border; c.fill = fill
                c.font = d_font
                c.alignment = center if col in (2, 3, 4, 5, 7, 8, 9) else left
            row_idx += 1

    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf); buf.seek(0)

    fname = f"kobber_variaciones_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ── Análisis completo + variaciones ───────────────────────────────────────────

@router.post("/full-report")
async def full_report(
    kobber_file: UploadFile = File(...),
    ml_file:     UploadFile = File(...),
):
    """
    Analiza el archivo ML contra la BD Y genera 3 variaciones de título por producto.
    Devuelve un Excel con dos hojas:
      - Resumen: estadísticas globales
      - Detalle: cada producto × 3 variaciones con estado y problemas marcados
    """
    from datetime import datetime
    from io import BytesIO

    kobber_bytes = await kobber_file.read()
    ml_bytes     = await ml_file.read()

    # ── Parsear archivos ───────────────────────────────────────────────────────
    try:
        kobber_index = _parse_kobber_file(kobber_bytes)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Error leyendo archivo Kobber: {e}")

    try:
        ml_rows = _parse_ml_file(ml_bytes)
    except Exception as e:
        raise HTTPException(400, f"Error leyendo archivo ML: {e}")

    # ── Cargar BD ──────────────────────────────────────────────────────────────
    db = get_client()
    bd_variants: dict = {}
    for p in db.table("products").select(
        "nombre, descripcion, product_variants(id, clave, codigo, nc, precio_distribuidor, stock)"
    ).execute().data:
        for v in (p.get("product_variants") or []):
            info = {
                "nombre":      p.get("nombre", ""),
                "descripcion": p.get("descripcion", ""),
                "nc":          str(v.get("nc") or "").strip(),
                "precio_dist": v.get("precio_distribuidor"),
            }
            clave  = str(v.get("clave")  or "").strip().lower()
            codigo = str(v.get("codigo") or "").strip().lower()
            if clave:  bd_variants[clave]  = info
            if codigo: bd_variants[codigo] = info

    # ── Análisis ───────────────────────────────────────────────────────────────
    analyzed = []
    claves_en_ml = set()
    for row in ml_rows:
        matched_clave, bd_info = _find_match(row, kobber_index, bd_variants)
        if matched_clave:
            claves_en_ml.add(matched_clave)

        issues = []
        if not matched_clave:
            issues.append(f"No encontrado en BD (modelo: {row['modelo']})")
        else:
            if row["precio"] is None:
                issues.append("Sin precio en ML")
            ean = row["ean"]
            nc_bd = bd_info.get("nc", "") if bd_info else ""
            ean_invalido = not ean or ean.lower() in ("escribe o elige un valor", "")
            if ean_invalido:
                if nc_bd and nc_bd not in ("", "2"):
                    issues.append(f"BD tiene NC={nc_bd} — falta EAN")
                else:
                    issues.append("Sin EAN ni NC útil en BD")
            if len(row["titulo"]) > 60:
                issues.append(f"Título largo ({len(row['titulo'])} chars)")
            if not row["codigo_ml"]:
                issues.append("Sin código catálogo ML (publicación libre)")
            if row["stock"] is None:
                issues.append("Sin stock")

        analyzed.append({
            "row":      row,
            "bd_info":  bd_info,
            "matched":  matched_clave,
            "issues":   issues,
        })

    no_en_output = [
        {"clave": e["clave"], "nombre": e["nombre"]}
        for clave_l, e in kobber_index.items()
        if clave_l not in claves_en_ml
    ]

    # ── Generar variaciones con Claude ─────────────────────────────────────────
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    BATCH = 10
    all_variations: list[list[str]] = []
    for i in range(0, len(ml_rows), BATCH):
        batch = ml_rows[i:i + BATCH]
        try:
            all_variations.extend(_generate_variations_batch(client, batch))
        except Exception:
            all_variations.extend([[r["titulo"][:60]] * 3 for r in batch])

    # ── Construir Excel ────────────────────────────────────────────────────────
    COPPER = "C8762C"; WHITE = "FFFFFF"
    thin   = Side(style="thin", color="D4CEC5")
    medium = Side(style="medium", color="D4CEC5")
    b_data = Border(left=thin,   right=thin,   top=thin,   bottom=thin)
    b_head = Border(left=medium, right=medium, top=medium, bottom=medium)

    h_font   = Font(bold=True, color=WHITE, size=10, name="Calibri")
    d_font   = Font(color="1A1510",         size=10, name="Calibri")
    ok_fill  = PatternFill("solid", fgColor="E8F5E9")   # verde suave
    warn_fill= PatternFill("solid", fgColor="FFF8E1")   # amarillo suave
    err_fill = PatternFill("solid", fgColor="FFEBEE")   # rojo suave
    alt_ok   = PatternFill("solid", fgColor="D4EDDA")
    alt_warn = PatternFill("solid", fgColor="FFF3CD")
    alt_err  = PatternFill("solid", fgColor="F8D7DA")
    h_fill   = PatternFill("solid", fgColor=COPPER)
    center   = Alignment(horizontal="center", vertical="center")
    left     = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen ────────────────────────────────────────────────────────
    ws_res = wb.active
    ws_res.title = "Resumen"

    total   = len(analyzed)
    con_iss = sum(1 for a in analyzed if a["issues"])
    sin_iss = total - con_iss

    resumen_rows = [
        ("Total filas en ML",        total),
        ("✅ Sin problemas",          sin_iss),
        ("⚠  Con problemas",         con_iss),
        ("👻 No aparecen en output", len(no_en_output)),
    ]
    issue_counts: dict = defaultdict(int)
    for a in analyzed:
        for iss in a["issues"]:
            key = iss.split("(")[0].strip()[:35]
            issue_counts[key] += 1

    ws_res.column_dimensions["A"].width = 38
    ws_res.column_dimensions["B"].width = 12

    ws_res.cell(1, 1, "RESUMEN DEL ANÁLISIS").font = Font(bold=True, size=12, name="Calibri")
    ws_res.cell(1, 1).fill = h_fill
    ws_res.cell(1, 1).font = Font(bold=True, color=WHITE, size=12, name="Calibri")
    ws_res.merge_cells("A1:B1")

    for i, (label, val) in enumerate(resumen_rows, 2):
        ws_res.cell(i, 1, label).font = d_font
        ws_res.cell(i, 2, val).font   = Font(bold=True, size=10, name="Calibri")
        ws_res.cell(i, 2).alignment   = center
        for c in [ws_res.cell(i,1), ws_res.cell(i,2)]:
            c.border = b_data

    ws_res.cell(7, 1, "Problemas encontrados").font = Font(bold=True, size=10, name="Calibri")
    ws_res.cell(7, 1).fill = PatternFill("solid", fgColor="F5F5F5")
    ws_res.merge_cells("A7:B7")

    for i, (label, cnt) in enumerate(sorted(issue_counts.items(), key=lambda x: -x[1]), 8):
        ws_res.cell(i, 1, label).font = d_font
        ws_res.cell(i, 2, cnt).font   = d_font
        ws_res.cell(i, 2).alignment   = center
        for c in [ws_res.cell(i,1), ws_res.cell(i,2)]:
            c.border = b_data

    # ── Hoja 2: Detalle ────────────────────────────────────────────────────────
    ws_det = wb.create_sheet("Detalle")
    det_headers = [
        ("Estado",          10), ("Var #",      6),  ("Título propuesto",  48),
        ("Título original", 48), ("Hoja ML",    20), ("Clave / Modelo",    16),
        ("Nombre en BD",    40), ("Marca",       14), ("Precio ML [$]",     14),
        ("EAN",             15), ("Stock",        8), ("Problemas",         55),
    ]
    ws_det.row_dimensions[1].height = 22
    ws_det.freeze_panes = "A2"

    for col, (h, w) in enumerate(det_headers, 1):
        c = ws_det.cell(1, col, h)
        c.font = h_font; c.fill = h_fill; c.border = b_head; c.alignment = center
        ws_det.column_dimensions[c.column_letter].width = w

    row_idx = 2
    for prod_idx, (entry, variations) in enumerate(zip(analyzed, all_variations)):
        row     = entry["row"]
        bd_info = entry["bd_info"]
        issues  = entry["issues"]

        # Elegir color según gravedad
        is_alt = prod_idx % 2 == 1
        if not issues:
            base_fill = alt_ok   if is_alt else ok_fill
            estado    = "✅ OK"
        elif any("No encontrado" in i or "Sin precio" in i for i in issues):
            base_fill = alt_err  if is_alt else err_fill
            estado    = "🔴 Error"
        else:
            base_fill = alt_warn if is_alt else warn_fill
            estado    = "⚠ Revisar"

        issues_str = " | ".join(issues) if issues else ""
        ean_val    = row["ean"] if row["ean"] not in ("", "Escribe o elige un valor") else ""

        for var_num, titulo in enumerate(variations, 1):
            data = [
                estado if var_num == 1 else "",
                var_num,
                titulo,
                row["titulo"] if var_num == 1 else "",
                row["hoja"]   if var_num == 1 else "",
                (row["modelo"] or row["sku"]) if var_num == 1 else "",
                (bd_info.get("nombre","") if bd_info else "") if var_num == 1 else "",
                row["marca"]  if var_num == 1 else "",
                row["precio"] if var_num == 1 else "",
                ean_val       if var_num == 1 else "",
                row["stock"]  if var_num == 1 else "",
                issues_str    if var_num == 1 else "",
            ]
            ws_det.row_dimensions[row_idx].height = 16
            for col, val in enumerate(data, 1):
                c = ws_det.cell(row_idx, col, val)
                c.border = b_data
                c.fill   = base_fill
                c.font   = d_font
                c.alignment = left if col in (3, 4, 7, 12) else center
            row_idx += 1

    buf = BytesIO()
    wb.save(buf); buf.seek(0)

    fname = f"kobber_analisis_ML_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ── Llenar archivo ML con datos de BD ─────────────────────────────────────────
#
# Columnas relevantes en el archivo ML (0-based → 1-based para openpyxl):
#   row[1]  → col 2  (B): Título          — se mantiene del archivo ML
#   row[4]  → col 5  (E): EAN             — se llena con NC de BD
#   row[6]  → col 7  (G): SKU             — se llena con clave de BD
#   row[7]  → col 8  (H): Stock           — se llena con stock de BD
#   row[8]  → col 9  (I): Precio          — se llena con precio_distribuidor de BD
#   row[9]  → col 10 (J): Descripción     — se llena con descripcion de BD
#   row[19] → col 20 (T): Marca           — se llena con marca de BD
#   row[20] → col 21 (U): Modelo          — se llena con clave de BD

@router.post("/generate-from-ml")
async def generate_from_ml(
    ml_file: UploadFile = File(...),
    margen:  float = 0,
):
    """
    Toma el archivo ML como plantilla, busca cada producto en la BD por modelo/SKU
    y llena las celdas (EAN, SKU, stock, precio, descripción, marca) con datos reales.
    Devuelve el mismo archivo con el formato original de ML, listo para subir.
    """
    from datetime import datetime
    from io import BytesIO

    ml_bytes = await ml_file.read()

    # Cargar el workbook preservando formato para modificarlo y devolverlo
    try:
        wb = openpyxl.load_workbook(io.BytesIO(ml_bytes))
    except Exception as e:
        raise HTTPException(400, f"Error leyendo archivo ML: {e}")

    # Cargar variantes de BD indexadas por clave y código
    db = get_client()
    bd_variants: dict = {}

    for p in db.table("products").select(
        "id, nombre, descripcion, marca, "
        "product_variants(id, clave, codigo, nc, precio_distribuidor, stock)"
    ).execute().data:
        for v in (p.get("product_variants") or []):
            info = {
                "descripcion": p.get("descripcion", ""),
                "marca":       p.get("marca", ""),
                "nc":          str(v.get("nc") or "").strip(),
                "precio_dist": v.get("precio_distribuidor"),
                "stock":       v.get("stock") or 0,
                "clave":       str(v.get("clave")  or "").strip(),
            }
            clave  = str(v.get("clave")  or "").strip().lower()
            codigo = str(v.get("codigo") or "").strip().lower()
            if clave:  bd_variants[clave]  = info
            if codigo: bd_variants[codigo] = info

    hojas = [s for s in wb.sheetnames if s.lower() not in ["ayuda", "extra info"]]

    total   = 0
    matches = 0

    for hoja in hojas:
        ws   = wb[hoja]
        cols = _ml_col_map(ws)   # posiciones reales de cada campo en esta hoja

        titulo_col = cols.get("titulo")
        if titulo_col is None:
            continue

        for row_num in range(9, ws.max_row + 1):
            titulo = str(ws.cell(row_num, titulo_col).value or "").strip()
            if not titulo:
                continue

            total += 1

            row_data = {
                "titulo": titulo,
                "modelo": str(ws.cell(row_num, cols["modelo"]).value or "").strip() if "modelo" in cols else "",
                "sku":    str(ws.cell(row_num, cols["sku"]).value    or "").strip() if "sku"    in cols else "",
            }

            _, bd = _find_match(row_data, {}, bd_variants)
            if not bd:
                continue

            matches += 1

            precio = bd["precio_dist"]
            if precio and margen:
                precio = round(precio * (1 + margen / 100), 2)

            # Es producto de catálogo ML si tiene código MCO... en col codigo_ml
            es_catalogo = bool(
                cols.get("codigo_ml") and
                str(ws.cell(row_num, cols["codigo_ml"]).value or "").strip()
            )

            # EAN / NC
            nc = bd["nc"]
            if nc and nc not in ("", "2") and "ean" in cols:
                ws.cell(row_num, cols["ean"]).value = nc

            # SKU / Clave
            if bd["clave"] and "sku" in cols:
                ws.cell(row_num, cols["sku"]).value = bd["clave"]

            # Stock
            if "stock" in cols:
                ws.cell(row_num, cols["stock"]).value = bd["stock"]

            # Para productos libres (sin código ML): también precio, descripción, marca, modelo
            if not es_catalogo:
                if precio is not None and "precio" in cols:
                    ws.cell(row_num, cols["precio"]).value = precio

                if bd["descripcion"] and "descripcion" in cols:
                    ws.cell(row_num, cols["descripcion"]).value = bd["descripcion"]

                if bd["marca"] and "marca" in cols:
                    ws.cell(row_num, cols["marca"]).value = bd["marca"]

                if bd["clave"] and "modelo" in cols:
                    ws.cell(row_num, cols["modelo"]).value = bd["clave"]

    buf = BytesIO()
    wb.save(buf); buf.seek(0)

    fname = f"kobber_ML_listo_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={fname}",
            "X-Total":   str(total),
            "X-Matches": str(matches),
        },
    )


# ── Descargar plantilla ML via scraper ────────────────────────────────────────

@router.post("/download-template")
async def download_template(body: dict):
    """
    Recibe una lista de product_ids, obtiene sus categorías ML de la BD,
    ejecuta el scraper headless y devuelve la plantilla descargada.
    """
    product_ids = body.get("product_ids", [])
    if not product_ids:
        raise HTTPException(400, "Se requieren product_ids")

    db = get_client()
    productos = db.table("products").select("nombre, categoria_ml") \
        .in_("id", product_ids).execute().data

    nombres = [p["nombre"] for p in productos if p.get("categoria_ml")]
    if not nombres:
        raise HTTPException(400, "Los productos seleccionados no tienen categoria_ml asignada")

    # Escribir lista de productos para el scraper
    nombres_file = "/tmp/ml_productos_seleccionados.txt"
    with open(nombres_file, "w") as f:
        f.write("\n".join(nombres))

    # Ejecutar el scraper como subprocess
    script = os.path.join(os.path.dirname(__file__), "../../scripts/ml_scrape_template.py")
    script = os.path.abspath(script)
    venv_python = os.path.join(os.path.dirname(__file__), "../../backend/venv/bin/python3")
    venv_python = os.path.abspath(venv_python)

    try:
        result = subprocess.run(
            [venv_python, script, "--file", nombres_file],
            capture_output=True, text=True, timeout=300,
            cwd=os.path.dirname(script),
        )
    except subprocess.TimeoutExpired:
        raise HTTPException(504, "El scraper tardó demasiado. Reintenta.")
    except Exception as e:
        raise HTTPException(500, f"Error ejecutando scraper: {e}")

    if result.returncode != 0:
        raise HTTPException(500, f"Error en el scraper:\n{result.stderr[-500:]}")

    # Buscar el archivo más reciente descargado
    downloads = sorted(
        glob.glob(os.path.expanduser("~/Downloads/Publicar-*.xlsx")),
        key=os.path.getmtime, reverse=True,
    )
    if not downloads:
        raise HTTPException(500, "El scraper no descargó ningún archivo")

    return FileResponse(
        downloads[0],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=os.path.basename(downloads[0]),
    )


# ── Rellenar plantilla vacía con productos de BD ──────────────────────────────

@router.post("/fill-blank-template")
async def fill_blank_template(
    ml_file:     UploadFile = File(...),
    margen:      float = 0,
    product_ids: Optional[str] = Form(None),  # JSON array de IDs o None = todos
):
    """
    Recibe una plantilla vacía de ML (descargada del scraper).
    Usa categoria_ml de la BD para saber qué productos van en cada hoja.
    Si se pasan product_ids, solo incluye esos productos.
    """
    from datetime import datetime
    from io import BytesIO

    ml_bytes = await ml_file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(ml_bytes))
    except Exception as e:
        raise HTTPException(400, f"Error leyendo archivo: {e}")

    hojas = [s for s in wb.sheetnames if s.lower() not in ["ayuda", "extra info"]]
    if not hojas:
        raise HTTPException(400, "No se encontraron hojas de categorías en el archivo")

    # Parsear product_ids si se enviaron
    ids_filter: list | None = None
    if product_ids:
        try:
            ids_filter = json.loads(product_ids)
        except Exception:
            ids_filter = [i.strip() for i in product_ids.split(",") if i.strip()]

    # ── Cargar BD con categoria_ml y atributos ────────────────────────────────
    db = get_client()
    query = db.table("products").select(
        "id, nombre, descripcion, marca, categoria_ml, caracteristicas, "
        "product_attributes(nombre, valor, unidad, variant_id), "
        "product_variants(id, clave, codigo, nc, precio_distribuidor, stock, "
        "product_attributes(nombre, valor, unidad)), "
        "product_images(url, orden)"
    ).not_.is_("categoria_ml", "null")

    if ids_filter:
        query = query.in_("id", ids_filter)

    productos_bd = query.execute().data

    # Mapa directo: categoria_ml → lista de productos
    cat_to_products: dict = {}
    for p in productos_bd:
        cat = p.get("categoria_ml")
        if cat:
            cat_to_products.setdefault(cat, []).append(p)

    # ── Detectar columnas de atributos en una hoja ─────────────────────────────
    _SKIP_COLS = {"buybox_formula", "hidden_pictures", "cantidad de caracteres",
                  "cargo por vender", "costo por ofrecer cuotas"}
    _STANDARD  = set(_ML_FIELD_KEYWORDS.keys())

    def _attr_col_maps(ws) -> tuple[dict, dict]:
        """
        Retorna (attr_cols, unit_cols):
          attr_cols: nombre_atributo_lower → col (columna de valor)
          unit_cols: nombre_atributo_lower → col (columna de unidad)
        """
        attr_cols: dict = {}
        unit_cols: dict = {}

        # Encontrar la fila de headers
        header_row = 3
        for row in range(2, 6):
            hits = sum(1 for c in range(1, ws.max_column + 1)
                       if ws.cell(row, c).value and
                       any(kw in str(ws.cell(row, c).value).lower()
                           for kws in _ML_FIELD_KEYWORDS.values() for kw in kws))
            if hits >= 4:
                header_row = row
                break

        for col in range(1, ws.max_column + 1):
            header = str(ws.cell(header_row, col).value or "").strip()
            if not header:
                continue
            header_l = header.lower()
            if any(skip in header_l for skip in _SKIP_COLS):
                continue
            # Columna de unidad: "Unidad de X"
            if header_l.startswith("unidad de "):
                attr_name = header_l[len("unidad de "):].strip()
                unit_cols[attr_name] = col
            # Columna de variante: "Varía por: X"
            elif header_l.startswith("varía por:") or header_l.startswith("varia por:"):
                attr_name = header_l.split(":", 1)[1].strip().split(" de ")[0].strip()
                attr_cols.setdefault(attr_name, col)
            # Columna de atributo normal (no estándar)
            elif not any(kw in header_l for kws in _ML_FIELD_KEYWORDS.values() for kw in kws):
                attr_cols[header_l] = col

        return attr_cols, unit_cols

    # ── Rellenar cada hoja ─────────────────────────────────────────────────────
    total_filas = 0

    for hoja in hojas:
        ws   = wb[hoja]
        cols = _ml_col_map(ws)
        if "titulo" not in cols:
            continue

        attr_cols, unit_cols = _attr_col_maps(ws)

        # Primera fila vacía de datos
        first_empty = 9
        while ws.cell(first_empty, cols["titulo"]).value:
            first_empty += 1

        # Copiar valores por defecto de la fila de ejemplo (fila 8)
        default_vals: dict = {}
        for field, col in cols.items():
            v = ws.cell(8, col).value
            if v and field not in ("titulo", "ean", "sku", "stock", "precio", "descripcion"):
                default_vals[col] = v

        productos_hoja = cat_to_products.get(hoja, [])

        for p in productos_hoja:
            fotos = sorted(p.get("product_images") or [], key=lambda x: x.get("orden", 0))
            # ML requiere URLs separadas por coma (ver hoja Ayuda del template)
            fotos_str = ",".join(f["url"] for f in fotos[:8])

            # Atributos de familia (variant_id = None)
            familia_attrs = {
                a["nombre"].lower(): a
                for a in (p.get("product_attributes") or [])
                if not a.get("variant_id")
            }

            # Título genérico para todas las variantes del producto (recomendación Ayuda ML)
            # Usar nombre + marca sin clave específica, para que sea igual en todas las variantes
            titulo_generico = f"{p.get('nombre', '')} {p.get('marca', '')}".strip()[:60]

            for v in (p.get("product_variants") or []):
                clave  = str(v.get("clave") or "").strip()
                nc     = str(v.get("nc") or "").strip()
                precio = v.get("precio_distribuidor")
                if precio and margen:
                    precio = round(precio * (1 + margen / 100), 2)

                # Atributos de variante
                variante_attrs = {
                    a["nombre"].lower(): a
                    for a in (v.get("product_attributes") or [])
                }
                # Fusionar: variante sobreescribe familia
                all_attrs = {**familia_attrs, **variante_attrs}

                # Escribir valores por defecto de la fila ejemplo
                for col, val in default_vals.items():
                    ws.cell(first_empty, col).value = val

                # Escribir datos de BD
                def w(field, val):
                    if field in cols and val not in (None, ""):
                        ws.cell(first_empty, cols[field]).value = val

                # Título igual para todas las variantes (según Ayuda ML)
                w("titulo",      titulo_generico)
                w("sku",         clave)
                w("modelo",      clave)
                w("marca",       p.get("marca", ""))
                w("descripcion", p.get("descripcion", ""))
                w("stock",       v.get("stock") or 0)
                w("condicion",   "Nuevo")
                if precio is not None:
                    w("precio", precio)
                if nc and nc not in ("", "2"):
                    w("ean", nc)
                # Fotos separadas por coma (según Ayuda ML)
                if "fotos" in cols and fotos_str:
                    ws.cell(first_empty, cols["fotos"]).value = fotos_str

                # ── Atributos específicos de la categoría ──────────────────
                for attr_name_l, attr in all_attrs.items():
                    valor  = str(attr.get("valor") or "").strip()
                    unidad = str(attr.get("unidad") or "").strip()

                    if not valor:
                        continue

                    # Buscar columna de valor por nombre exacto o parcial
                    target_col = attr_cols.get(attr_name_l)
                    if not target_col:
                        # Búsqueda parcial: el nombre del atributo aparece en el header
                        for col_name, col_num in attr_cols.items():
                            if attr_name_l in col_name or col_name in attr_name_l:
                                target_col = col_num
                                break

                    if target_col:
                        ws.cell(first_empty, target_col).value = valor
                        # Buscar columna de unidad asociada
                        if unidad:
                            unit_col = unit_cols.get(attr_name_l)
                            if not unit_col:
                                for col_name, col_num in unit_cols.items():
                                    if attr_name_l in col_name or col_name in attr_name_l:
                                        unit_col = col_num
                                        break
                            if unit_col:
                                ws.cell(first_empty, unit_col).value = unidad

                first_empty  += 1
                total_filas  += 1

    buf = BytesIO()
    wb.save(buf); buf.seek(0)

    fname = f"kobber_ML_completo_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={fname}",
            "X-Filas": str(total_filas),
            "X-Hojas": str(len(hojas)),
        },
    )
